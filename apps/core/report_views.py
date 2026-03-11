import io
import json
from datetime import timedelta

from django.contrib.auth.decorators import login_required
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render
from django.utils import timezone


def _get_report_data(date_from, date_to):
    from apps.alerts.models import Alert
    from apps.incidents.models import Incident
    from datetime import datetime

    tz = timezone.get_current_timezone()
    dt_from = timezone.make_aware(datetime.combine(date_from, datetime.min.time()), tz)
    dt_to   = timezone.make_aware(datetime.combine(date_to,   datetime.max.time().replace(microsecond=0)), tz)

    alerts = Alert.objects.filter(timestamp__gte=dt_from, timestamp__lte=dt_to)
    total  = alerts.count()
    sev_counts = {s: alerts.filter(severity=s).count() for s in ['CRITICAL','HIGH','MEDIUM','LOW','INFO']}

    top_rules = list(
        alerts.values('rule_id','rule_description')
        .annotate(count=Count('id')).order_by('-count')[:10]
    )
    top_agents = list(
        alerts.values('agent_name','agent_ip')
        .annotate(count=Count('id')).order_by('-count')[:10]
    )
    mitre = list(
        alerts.exclude(mitre_id='').exclude(mitre_id__isnull=True)
        .values('mitre_id').annotate(count=Count('id')).order_by('-count')[:10]
    )

    from apps.alerts.models import AIAnalysis, AIAnalysisChat
    analyzed      = AIAnalysis.objects.filter(alert__timestamp__gte=dt_from, alert__timestamp__lte=dt_to).count()
    analyzed_chat = AIAnalysisChat.objects.filter(alert__timestamp__gte=dt_from, alert__timestamp__lte=dt_to).count()
    coverage_pct  = round(max(analyzed, analyzed_chat) / total * 100) if total else 0

    incidents_new      = Incident.objects.filter(created_at__gte=dt_from, created_at__lte=dt_to).count()
    incidents_resolved = Incident.objects.filter(created_at__gte=dt_from, created_at__lte=dt_to, status='Resolved').count()
    incidents_open     = Incident.objects.filter(created_at__gte=dt_from, created_at__lte=dt_to).exclude(status__in=['Resolved','Closed']).count()

    return {
        'date_from': date_from, 'date_to': date_to,
        'total': total, 'sev_counts': sev_counts,
        'top_rules': top_rules, 'top_agents': top_agents, 'mitre': mitre,
        'coverage_pct': coverage_pct, 'analyzed': max(analyzed, analyzed_chat),
        'incidents_new': incidents_new, 'incidents_resolved': incidents_resolved,
        'incidents_open': incidents_open,
        'generated_at': timezone.localtime(timezone.now()),
    }


@login_required
def report_page(request):
    today    = timezone.localdate()
    week_ago = today - timedelta(days=6)
    return render(request, 'core/report.html', {
        'default_from': week_ago.strftime('%Y-%m-%d'),
        'default_to':   today.strftime('%Y-%m-%d'),
    })


@login_required
def report_preview(request):
    from datetime import date as date_type
    try:
        date_from = date_type.fromisoformat(request.GET.get('date_from',''))
        date_to   = date_type.fromisoformat(request.GET.get('date_to',''))
    except ValueError:
        return HttpResponse('Invalid date', status=400)
    data = _get_report_data(date_from, date_to)
    data['date_from']    = str(data['date_from'])
    data['date_to']      = str(data['date_to'])
    data['generated_at'] = data['generated_at'].strftime('%Y-%m-%d %H:%M')
    return HttpResponse(json.dumps(data, ensure_ascii=False), content_type='application/json')


@login_required
def report_excel(request):
    from datetime import date as date_type
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    try:
        date_from = date_type.fromisoformat(request.GET.get('date_from',''))
        date_to   = date_type.fromisoformat(request.GET.get('date_to',''))
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    data = _get_report_data(date_from, date_to)
    wb   = openpyxl.Workbook()

    hdr_fill = PatternFill('solid', fgColor='1f6feb')
    hdr_font = Font(bold=True, color='FFFFFF')
    def thin(): s = Side(style='thin', color='CCCCCC'); return Border(left=s,right=s,top=s,bottom=s)

    # Sheet 1: Summary
    ws = wb.active
    ws.title = 'Summary'
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 18
    ws['A1'] = 'SOC Security Report'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A2'] = f'Period: {date_from} to {date_to}'
    ws['A3'] = f'Generated: {data["generated_at"].strftime("%Y-%m-%d %H:%M")}'
    ws.append([])
    ws.append(['Metric', 'Count'])
    for cell in ws[5]: cell.fill = hdr_fill; cell.font = hdr_font
    rows = [
        ('Total Alerts', data['total']),
        ('Critical', data['sev_counts']['CRITICAL']),
        ('High',     data['sev_counts']['HIGH']),
        ('Medium',   data['sev_counts']['MEDIUM']),
        ('Low',      data['sev_counts']['LOW']),
        ('Info',     data['sev_counts']['INFO']),
        ('', ''),
        ('Incidents (period)', data['incidents_new']),
        ('Incidents Open',     data['incidents_open']),
        ('Incidents Resolved', data['incidents_resolved']),
        ('', ''),
        ('AI Coverage', f'{data["coverage_pct"]}%'),
    ]
    sev_colors = {'Critical':'FFCCCC','High':'FFE5B4','Medium':'CCE5FF','Low':'D4EDDA'}
    for r in rows:
        ws.append(list(r))
    for row in ws.iter_rows(min_row=6, max_row=ws.max_row):
        label = str(row[0].value)
        if label in sev_colors:
            for cell in row: cell.fill = PatternFill('solid', fgColor=sev_colors[label])
        for cell in row: cell.border = thin()

    # Sheet 2: Top Rules
    ws2 = wb.create_sheet('Top Rules')
    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 50
    ws2.column_dimensions['C'].width = 12
    ws2.append(['Rule ID','Description','Count'])
    for cell in ws2[1]: cell.fill = PatternFill('solid',fgColor='238636'); cell.font = hdr_font
    for r in data['top_rules']:
        ws2.append([r['rule_id'], r['rule_description'] or '', r['count']])
    for row in ws2.iter_rows(min_row=2):
        for cell in row: cell.border = thin()

    # Sheet 3: Top Agents
    ws3 = wb.create_sheet('Top Agents')
    ws3.column_dimensions['A'].width = 22
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 12
    ws3.append(['Agent Name','Agent IP','Count'])
    for cell in ws3[1]: cell.fill = PatternFill('solid',fgColor='f0883e'); cell.font = hdr_font
    for a in data['top_agents']:
        ws3.append([a['agent_name'] or '', a['agent_ip'] or '', a['count']])
    for row in ws3.iter_rows(min_row=2):
        for cell in row: cell.border = thin()

    # Sheet 4: MITRE
    ws4 = wb.create_sheet('MITRE ATT&CK')
    ws4.column_dimensions['A'].width = 20
    ws4.column_dimensions['B'].width = 12
    ws4.append(['Technique ID','Count'])
    for cell in ws4[1]: cell.fill = PatternFill('solid',fgColor='8957e5'); cell.font = hdr_font
    for m in data['mitre']:
        ws4.append([m['mitre_id'], m['count']])
    for row in ws4.iter_rows(min_row=2):
        for cell in row: cell.border = thin()

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    filename = f'SOC_Report_{date_from}_{date_to}.xlsx'
    resp = HttpResponse(buf.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp


@login_required
def report_pdf(request):
    from datetime import date as date_type
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import cm
    from reportlab.lib import colors
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.enums import TA_CENTER

    try:
        date_from = date_type.fromisoformat(request.GET.get('date_from',''))
        date_to   = date_type.fromisoformat(request.GET.get('date_to',''))
    except ValueError:
        return HttpResponse('Invalid date', status=400)

    data   = _get_report_data(date_from, date_to)
    buf    = io.BytesIO()
    doc    = SimpleDocTemplate(buf, pagesize=A4,
                rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    story  = []

    title_s = ParagraphStyle('t', parent=styles['Title'], fontSize=18, alignment=TA_CENTER)
    sub_s   = ParagraphStyle('s', parent=styles['Normal'], fontSize=10, textColor=colors.grey, alignment=TA_CENTER)
    h2_s    = ParagraphStyle('h', parent=styles['Heading2'], fontSize=12, textColor=colors.HexColor('#1f6feb'), spaceBefore=12)

    story += [
        Paragraph('SOC Security Report', title_s),
        Paragraph(f'Period: {date_from} to {date_to}', sub_s),
        Paragraph(f'Generated: {data["generated_at"].strftime("%Y-%m-%d %H:%M")}', sub_s),
        HRFlowable(width='100%', thickness=1, color=colors.lightgrey),
        Spacer(1, 0.3*cm),
    ]

    # Summary table
    story.append(Paragraph('Alert Summary', h2_s))
    td = [['Metric','Count'],
          ['Critical', str(data['sev_counts']['CRITICAL'])],
          ['High',     str(data['sev_counts']['HIGH'])],
          ['Medium',   str(data['sev_counts']['MEDIUM'])],
          ['Low',      str(data['sev_counts']['LOW'])],
          ['Total',    str(data['total'])]]
    t = Table(td, colWidths=[10*cm, 5*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#1f6feb')),
        ('TEXTCOLOR', (0,0),(-1,0),colors.white),
        ('FONTNAME',  (0,0),(-1,0),'Helvetica-Bold'),
        ('BACKGROUND',(0,1),(-1,1),colors.HexColor('#FFCCCC')),
        ('BACKGROUND',(0,2),(-1,2),colors.HexColor('#FFE5B4')),
        ('BACKGROUND',(0,3),(-1,3),colors.HexColor('#CCE5FF')),
        ('BACKGROUND',(0,4),(-1,4),colors.HexColor('#D4EDDA')),
        ('BACKGROUND',(0,5),(-1,5),colors.HexColor('#EEEEEE')),
        ('FONTNAME',  (0,5),(-1,5),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('ALIGN',(1,0),(1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),5),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    story += [t, Spacer(1, 0.2*cm)]

    # Incidents
    story.append(Paragraph('Incidents', h2_s))
    id_ = [['Status','Count'],
           ['Total (period)', str(data['incidents_new'])],
           ['Open/InProgress', str(data['incidents_open'])],
           ['Resolved',        str(data['incidents_resolved'])]]
    it = Table(id_, colWidths=[10*cm, 5*cm])
    it.setStyle(TableStyle([
        ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#136e6e')),
        ('TEXTCOLOR', (0,0),(-1,0),colors.white),
        ('FONTNAME',  (0,0),(-1,0),'Helvetica-Bold'),
        ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
        ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F8F9FA')]),
        ('ALIGN',(1,0),(1,-1),'CENTER'),
        ('VALIGN',(0,0),(-1,-1),'MIDDLE'),
        ('TOPPADDING',(0,0),(-1,-1),5),
        ('BOTTOMPADDING',(0,0),(-1,-1),5),
    ]))
    story += [it, Paragraph(f'AI Coverage: {data["coverage_pct"]}%', styles['Normal']), Spacer(1,0.2*cm)]

    # Top Rules
    if data['top_rules']:
        story.append(Paragraph('Top Alert Rules', h2_s))
        rd = [['#','Rule ID','Description','Count']]
        for i,r in enumerate(data['top_rules'],1):
            desc = (r['rule_description'] or '')[:45] + ('...' if len(r['rule_description'] or '')>45 else '')
            rd.append([str(i), str(r['rule_id']), desc, str(r['count'])])
        rt = Table(rd, colWidths=[0.8*cm,2.5*cm,11*cm,2*cm])
        rt.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#238636')),
            ('TEXTCOLOR', (0,0),(-1,0),colors.white),
            ('FONTNAME',  (0,0),(-1,0),'Helvetica-Bold'),
            ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F8F9FA')]),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story += [rt, Spacer(1, 0.2*cm)]

    # Top Agents
    if data['top_agents']:
        story.append(Paragraph('Top Attacked Agents', h2_s))
        ad = [['#','Agent Name','Agent IP','Count']]
        for i,a in enumerate(data['top_agents'],1):
            ad.append([str(i), a['agent_name'] or '-', a['agent_ip'] or '-', str(a['count'])])
        at = Table(ad, colWidths=[0.8*cm,8*cm,5*cm,2.5*cm])
        at.setStyle(TableStyle([
            ('BACKGROUND',(0,0),(-1,0),colors.HexColor('#f0883e')),
            ('TEXTCOLOR', (0,0),(-1,0),colors.white),
            ('FONTNAME',  (0,0),(-1,0),'Helvetica-Bold'),
            ('GRID',(0,0),(-1,-1),0.5,colors.lightgrey),
            ('ROWBACKGROUNDS',(0,1),(-1,-1),[colors.white,colors.HexColor('#F8F9FA')]),
            ('FONTSIZE',(0,0),(-1,-1),8),
            ('TOPPADDING',(0,0),(-1,-1),4),('BOTTOMPADDING',(0,0),(-1,-1),4),
        ]))
        story.append(at)

    doc.build(story)
    buf.seek(0)
    filename = f'SOC_Report_{date_from}_{date_to}.pdf'
    resp = HttpResponse(buf.read(), content_type='application/pdf')
    resp['Content-Disposition'] = f'attachment; filename="{filename}"'
    return resp
